import openpyxl
import json
import os
import argparse
from .gcs_storage import get_gcs_manager

def find_col_index(sheet, year):
    """Finds the column index for a given year in the header row."""
    # Convert input year to string for comparison
    year_str = str(year).strip()
    
    for row in sheet.iter_rows(min_row=9, max_row=9):
        for cell in row:
            if cell.value:
                # Convert cell value to string for robust comparison
                cell_val_str = str(cell.value).strip()
                if cell_val_str == year_str:
                    return cell.column
                # Also try partial matching if exact match fails (e.g. "FY 2024-25" vs "2024-25")
                if year_str in cell_val_str or cell_val_str in year_str:
                     # Be careful with partial matches, but for FY it might be needed if template differs
                     pass
                     
    return None

def update_excel(data, excel_template_path, output_path, years_list, config_path=None):
    """
    Updates an Excel template with data from a JSON file for multiple years.
    
    Args:
        data: Multi-year data dict {year: {field: data}}
        excel_template_path: Path to Excel template
        output_path: Path for output file
        years_list: List of years to process
        config_path: Optional path to a local fields config file.
    """
    fields_config_path = None
    gcs_manager = get_gcs_manager()

    try:
        if config_path:
            print(f"Loading fields configuration from local path: {config_path}")
            fields_config_path = config_path
        elif gcs_manager:
            try:
                print("Attempting to load fields configuration from GCS...")
                fields_config_path = gcs_manager.download_fields_config()
                print("Successfully loaded configuration from GCS.")
            except Exception as gcs_error:
                print(f"GCS config download failed: {gcs_error}")
                print("Falling back to local configuration.")
                fields_config_path = os.path.join('prompts', 'fieldstoextract.json')
        else:
            print("GCS not available. Using local configuration.")
            fields_config_path = os.path.join('prompts', 'fieldstoextract.json')

        with open(fields_config_path, 'r') as f:
            config = json.load(f)
        
        mapping = {item['fieldName']: {'rowID': item['rowID'], 'sheetName': item['sheetName']} for item in config['fieldsToExtract']}

        workbook = openpyxl.load_workbook(excel_template_path)
        print(f"Loaded new workbook from template: {excel_template_path}")

        reference_sheet = workbook['Form-II']
        year_columns = {}
        
        for year in years_list:
            col_index = find_col_index(reference_sheet, year)
            if not col_index:
                print(f"Error: Year '{year}' not found in the reference sheet 'Form-II'.")
                continue
            year_columns[year] = col_index
            print(f"Found year {year} at column {col_index}")
        
        if not year_columns:
            print("Error: No valid years found in the spreadsheet.")
            return
        
        for year in years_list:
            if year not in year_columns:
                continue
                
            col_index = year_columns[year]
            year_data = data.get(str(year), {})
            
            print(f"Processing {len(year_data)} fields for year {year}")
            
            for field, value_info in year_data.items():
                if field in mapping:
                    map_info = mapping[field]
                    sheet_name = map_info['sheetName']
                    row_index = map_info['rowID']
                    
                    if row_index == 0:
                        continue

                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        target_cell = sheet.cell(row=row_index, column=col_index)
                        value_to_update = value_info[0]['value'] if value_info else ""
                        
                        if value_to_update in [None, "None", "null", "N/A", "n/a", "NULL"]:
                            value_to_update = ""
                        
                        if isinstance(value_to_update, str) and value_to_update.strip().lower() in ["none", "null", "n/a"]:
                            value_to_update = ""
                        
                        if value_to_update != "":
                            try:
                                if '.' in str(value_to_update):
                                    value_to_update = float(value_to_update)
                                else:
                                    value_to_update = int(value_to_update)
                            except (ValueError, TypeError):
                                value_to_update = str(value_to_update)
                        
                        if target_cell.value and isinstance(target_cell.value, str) and target_cell.value.startswith('='):
                            continue
                        
                        sheet.cell(row=row_index, column=col_index, value=value_to_update)
                    else:
                        print(f"Warning: Sheet '{sheet_name}' not found in the workbook.")
        
        # Save the workbook directly to the output path.
        workbook.save(output_path)
        print(f"\nSuccessfully updated Excel file saved to: {output_path}")

    except FileNotFoundError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        if 'gcs' in locals() and 'fields_config_path' in locals() and os.path.exists(fields_config_path) and not config_path:
            os.unlink(fields_config_path)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Update an Excel template with JSON data.")
    parser.add_argument("json_path", help="Path to the input JSON data file.")
    parser.add_argument("excel_template_path", help="Path to the Excel template file.")
    parser.add_argument("output_path", help="Path to save the updated Excel file.")
    parser.add_argument("--years", help="Comma-separated list of years to process for multi-year data.")
    parser.add_argument("--config", help="Path to a local fields configuration file (optional).")
    
    args = parser.parse_args()
    
    try:
        with open(args.json_path, 'r') as f:
            data = json.load(f)
        
        years_list = [year.strip() for year in args.years.split(',')] if args.years else None
        
        update_excel(data, args.excel_template_path, args.output_path, years_list=years_list, config_path=args.config)
        
    except FileNotFoundError:
        print(f"Error: The file '{args.json_path}' was not found.")
    except json.JSONDecodeError:
        print(f"Error: The file '{args.json_path}' is not a valid JSON file.")
    except Exception as e:
        print(f"An error occurred: {e}")
